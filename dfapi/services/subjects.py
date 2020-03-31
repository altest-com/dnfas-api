from os import path
from typing import List
from typing import Tuple
import datetime

import numpy as np
from django.db.models import QuerySet
from django.http import QueryDict, HttpRequest
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.cell import Cell

from ..models import Subject
from ..models import (
    SubjectSegment
)


def build_queryset(
    queryset: QuerySet,
    params: QueryDict
) -> Tuple[QuerySet, bool]:

    filtered = False

    name = params.get('name', None)
    if name is not None:
        queryset = queryset.filter(name__icontains=name)
        filtered = True

    last_name = params.get('last_name', None)
    if last_name is not None:
        queryset = queryset.filter(last_name__icontains=last_name)
        filtered = True

    naming = params.get('naming', None)
    if naming == SubjectSegment.NAMING_NAMED:
        queryset = queryset.exclude(
            name='', last_name=''
        )
        filtered = True
    elif naming == SubjectSegment.NAMING_UNNAMED:
        queryset = queryset.filter(
            name='', last_name=''
        )
        filtered = True

    tasks = params.getlist('tasks', None)
    if tasks is not None and len(tasks):
        queryset = queryset.filter(faces__task__in=tasks)
        filtered = True

    tasks_tags = params.getlist('tasks_tags', None)
    if tasks_tags is not None and len(tasks_tags):
        queryset = queryset.filter(faces__task__tags__in=tasks_tags)
        filtered = True

    min_time = params.get('min_time', None)
    if min_time is not None:
        queryset = queryset.filter(faces__created_at__time__gte=min_time)
        filtered = True

    max_time = params.get('max_time', None)
    if max_time is not None:
        queryset = queryset.filter(faces__created_at__time__lte=max_time)
        filtered = True

    min_date = params.get('min_date', None)
    if min_date is not None:
        queryset = queryset.filter(faces__created_at__date__gte=min_date)
        filtered = True

    max_date = params.get('max_date', None)
    if max_date is not None:
        queryset = queryset.filter(faces__created_at__date__lte=max_date)
        filtered = True

    max_age = params.get('max_age', None)
    if max_age is not None:
        try:
            max_age = int(max_age)
        except ValueError:
            pass
        else:
            queryset = queryset.filter(
                birthdate__gt=Subject.birthdate_from_age(max_age)
            )
            filtered = True

    min_age = params.get('min_age', None)
    if min_age is not None:
        try:
            min_age = int(min_age)
        except ValueError:
            pass
        else:
            queryset = queryset.filter(
                birthdate__lt=Subject.birthdate_from_age(min_age)
            )
            filtered = True

    min_pred_age = params.get('min_pred_age', None)
    if min_pred_age is not None:
        try:
            min_pred_age = int(min_pred_age)
        except ValueError:
            pass
        else:
            queryset = queryset.filter(pred_age__gte=min_pred_age)
            filtered = True

    max_pred_age = params.get('max_pred_age', None)
    if max_pred_age is not None:
        try:
            max_pred_age = int(max_pred_age)
        except ValueError:
            pass
        else:
            queryset = queryset.filter(pred_age__lte=max_pred_age)
            filtered = True

    sex = params.get('sex', None)
    if sex is not None:
        queryset = queryset.filter(sex=sex)
        filtered = True

    pred_sex = params.get('pred_sex', None)
    if pred_sex is not None:
        queryset = queryset.filter(pred_sex=pred_sex)
        filtered = True

    skin = params.get('skin', None)
    if skin is not None:
        queryset = queryset.filter(skin=skin)
        filtered = True

    order_by = params.get('order_by', None)
    if order_by is not None:
        queryset = queryset.order_by(order_by)

    if filtered:
        queryset = queryset.distinct()

    return queryset, filtered


def pred_sexage(subject: Subject):
    faces = subject.faces.all()
    age = None
    age_var = 0
    sex = ''
    sex_score = 0
    ages_sum = 0
    ages_var_sum = 0
    ages_count = 0
    man_count = 0
    woman_count = 0
    sex_scores_man_sum = 0
    sex_scores_woman_sum = 0

    for face in faces:
        if face.pred_age:
            ages_sum += face.pred_age
            ages_count += 1
            ages_var_sum += face.pred_age_var

        pred_sex = face.pred_sex
        if pred_sex:
            if pred_sex == subject.SEX_MAN:
                man_count += 1
                sex_scores_man_sum += face.pred_sex_score
            elif pred_sex == subject.SEX_WOMAN:
                woman_count += 1
                sex_scores_woman_sum += face.pred_sex_score

    if ages_count:
        age = ages_sum / ages_count
        age_var = ages_var_sum / ages_count

    if man_count + woman_count:
        if man_count > woman_count:
            sex = subject.SEX_MAN
            sex_score = sex_scores_man_sum / man_count
        else:
            sex = subject.SEX_WOMAN
            sex_score = sex_scores_woman_sum / woman_count

    return age, age_var, sex, sex_score


def demograp(subjects_queryset: QuerySet):
    men_ages = []
    women_ages = []
    men_count = 0
    women_count = 0

    for subject in subjects_queryset.iterator():
        age = subject.pred_age
        sex = subject.pred_sex
        if age and sex:
            if sex == Subject.SEX_MAN:
                men_ages.append(age)
                men_count += 1
            elif sex == Subject.SEX_WOMAN:
                women_ages.append(age)
                women_count += 1

    lower = 18
    upper = 74
    step = 6
    labels = list(range(lower, upper, step))

    return {
        'age_labels': labels,
        'men_ages': _age_stats(men_ages, labels),
        'women_ages': _age_stats(women_ages, labels),
        'men_count': men_count,
        'women_count': women_count,
        'hourly_count': _hourly_count(subjects_queryset),
        'daily_count': _daily_count(subjects_queryset),
    }


def _hourly_count(subjects_queryset: QuerySet):

    hours = [str(datetime.time(hour=hour)) for hour in range(0, 24)]
    men_count = []
    women_count = []

    if subjects_queryset.count():
        for ind in range(len(hours) - 1):
            hour_lower = hours[ind]
            hour_upper = hours[ind + 1]
            subjects = subjects_queryset.filter(
                faces__created_at__time__gte=hour_lower,
                faces__created_at__time__lt=hour_upper
            )
            men_count.append(len(set([
                subject.id for subject in subjects
                if subject.pred_sex == Subject.SEX_MAN
            ])))
            women_count.append(len(set([
                subject.id for subject in subjects
                if subject.pred_sex == Subject.SEX_WOMAN
            ])))

    return {
        'hours': hours,
        'men_count': men_count,
        'women_count': women_count,
    }


def _daily_count(subjects_queryset: QuerySet):


    dates = []
    men_count = []
    women_count = []

    if subjects_queryset.count():
        subjects_queryset = subjects_queryset.order_by('faces__created_at')
        oldest_subject = subjects_queryset.first()
        newest_subject = subjects_queryset.last()

        lower_timestamp = min([
            face.created_at for face in oldest_subject.faces.all()
        ])
        upper_timestamp = max([
            face.created_at for face in newest_subject.faces.all()
        ])

        lower_date = timezone.localtime(lower_timestamp).date()
        upper_date = timezone.localtime(upper_timestamp).date()

        curr_date = lower_date
        ind = 0

        while curr_date <= upper_date:
            next_date = curr_date + datetime.timedelta(days=1)

            subjects = subjects_queryset.filter(
                faces__created_at__date__gte=curr_date,
                faces__created_at__date__lt=next_date
            )
            men_count.append(len(set([
                subject.id for subject in subjects
                if subject.pred_sex == Subject.SEX_MAN
            ])))
            women_count.append(len(set([
                subject.id for subject in subjects
                if subject.pred_sex == Subject.SEX_WOMAN
            ])))

            dates.append(str(curr_date))
            curr_date = next_date
            ind += 1

    return {
        'dates': dates,
        'men_count': men_count,
        'women_count': women_count,
    }


def _age_stats(ages, labels):
    if not len(ages):
        return {
            'counts': [0] * len(labels),
            'mean_value': 0,
            'min_value': 0,
            'max_value': 0
        }

    mean_age = np.mean(ages)
    max_age = np.max(ages)
    min_age = np.min(ages)

    limits = (
        min(0, labels[0] - 1),
        max(max_age, labels[len(labels) - 1])
    )
    counts, _ = np.histogram(ages, bins=([limits[0]] + labels + [limits[1]]))

    return {
        'counts': counts,
        'mean_value': mean_age,
        'min_value': min_age,
        'max_value': max_age
    }


def _fill_xls_cell(subject: Subject, cell: Cell, col_key: str, request: HttpRequest):
    if col_key == 'id':
        cell.value = subject.pk
    elif col_key == 'image':
        image = subject.image
        if image is None:
            cell.value = ''
        else:
            cell.value = path.basename(image.path)
            cell.hyperlink = request.build_absolute_uri(image.url)
    elif col_key == 'created_at':
        cell.value = subject.created_at
    elif col_key == 'pred_sex':
        cell.value = subject.pred_sex
    elif col_key == 'pred_age':
        cell.value = subject.pred_age
    else:
        raise ValueError(f'Invalid column "{col_key}".')


def xls_export(
    queryset: QuerySet,
    request: HttpRequest,
    title: str = 'Subjects',
    columns: List[str] = None
):

    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = title

    fields = {
        'id': 'Id',
        'image': 'Image',
        'created_at': 'Created at',
        'pred_sex': 'Predicted sex',
        'pred_age': 'Predicted age'
    }

    if columns is not None:
        fields = {key: val for key, val in fields.items() if key in columns}

    # Define the titles for fields
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, col_key in enumerate(fields.keys(), 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = fields[col_key]

    # Iterate through all movies
    for subject in queryset:
        row_num += 1
        # Assign the data for each cell of the row
        for col_num, col_key in enumerate(fields, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            _fill_xls_cell(subject, cell, col_key, request)

    return workbook